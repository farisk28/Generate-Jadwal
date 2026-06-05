import streamlit as st
import pandas as pd
import numpy as np
from ortools.sat.python import cp_model
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import datetime

# Setup judul aplikasi
st.set_page_config(page_title="Automated Scheduling System", layout="wide")
st.title("🗓️ Sistem Otomasi Penjadwalan Kerja")
st.write("Upload jadwal bulan lalu untuk membuat jadwal bulan berikutnya secara otomatis sesuai aturan.")

# =====================================================================
# FUNCTION: PARSING JADWAL BULAN LALU (OTOMATIS AMBIL 7 HARI TERAKHIR)
# =====================================================================
def baca_riwayat_bulan_lalu(uploaded_file):
    df = pd.read_excel(uploaded_file, header=None)
    karyawan_list = []
    riwayat_weekly = {}
    
    for idx, row in df.iterrows():
        if idx >= 3 and pd.notna(row[0]) and "Monitoring" in str(row[0]):
            nama_emp = str(row[0]).strip()
            karyawan_list.append(nama_emp)
            
            columns_with_data = []
            for col_idx in range(8, df.shape[1]):
                val_tgl = df.iloc[1, col_idx]
                if pd.notna(val_tgl) and str(val_tgl).isdigit():
                    columns_with_data.append(col_idx)
            
            kolom_7_hari_terakhir = columns_with_data[-7:]
            
            raw_shifts = []
            for col in kolom_7_hari_terakhir:
                val_shift = str(df.iloc[idx, col]).strip()
                if val_shift == '1': raw_shifts.append(1)
                elif val_shift == '2': raw_shifts.append(2)
                elif val_shift == '3': raw_shifts.append(3)
                else: raw_shifts.append(0)
                
            riwayat_weekly[nama_emp] = raw_shifts
            
    bulan_lalu_raw = str(df.iloc[0, 0]).strip()
    if " " in bulan_lalu_raw:
        bulan_lalu_raw = bulan_lalu_raw.split(" ")[0]
        
    date_bulan_lalu = datetime.datetime.strptime(bulan_lalu_raw, "%Y-%m-%d")
    
    return karyawan_list, riwayat_weekly, date_bulan_lalu

# =====================================================================
# INTERFACE STREAMLIT STEP 1 & 2: UPLOAD FILE
# =====================================================================
uploaded_file = st.file_uploader("Pilih file Excel Jadwal Bulan Lalu (.xlsx)", type=["xlsx"])

if uploaded_file is not None:
    try:
        karyawan, riwayat_weekly_dict, date_bulan_lalu = baca_riwayat_bulan_lalu(uploaded_file)
        num_karyawan = len(karyawan)
        
        riwayat_mei_seminggu = [riwayat_weekly_dict[name] for name in karyawan]
        
        if date_bulan_lalu.month == 12:
            thn_depan = date_bulan_lalu.year + 1
            bln_depan = 1
        else:
            thn_depan = date_bulan_lalu.year
            bln_depan = date_bulan_lalu.month + 1
            
        date_bulan_depan = datetime.date(thn_depan, bln_depan, 1)
        
        if bln_depan in [1, 3, 5, 7, 8, 10, 12]: num_hari = 31
        elif bln_depan in [4, 6, 9, 11]: num_hari = 30
        else: num_hari = 29 if thn_depan % 4 == 0 else 28
        
        hari_pertama_idx = date_bulan_depan.weekday()
        nama_hari_format = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
        hari_ke_nama = [nama_hari_format[(hari_pertama_idx + d) % 7] for d in range(num_hari)]
        
        st.success(f"Berhasil membaca data! Menyiapkan jadwal otomatis untuk Bulan: **{date_bulan_depan.strftime('%B %Y')}** ({num_hari} Hari).")
        
        with st.expander("Lihat Karyawan Terdeteksi & Riwayat Akhir Bulan"):
            for i, name in enumerate(karyawan):
                st.write(f"- {name} | Riwayat 7 hari terakhir: {riwayat_mei_seminggu[i]}")

        # =====================================================================
        # STEP 3: TOMBOL GENERATE JADWAL
        # =====================================================================
        if st.button("🚀 Jalankan Otomasi Penjadwalan"):
            with st.spinner("Algoritma sedang menghitung kombinasi terbaik... Mohon tunggu."):
                model = cp_model.CpModel()
                x = {}

                saut_idx = -1
                for idx, name in enumerate(karyawan):
                    if "Saut" in name:
                        saut_idx = idx
                        break

                for i in range(num_karyawan):
                    for d in range(num_hari):
                        if i == 0: # Jasmine hanya boleh Shift 1, Shift 2, atau OFF
                            x[i, d] = model.NewIntVarFromDomain(cp_model.Domain.FromValues([0, 1, 2]), f'x_{i}_{d}')
                        else:
                            x[i, d] = model.NewIntVarFromDomain(cp_model.Domain.FromValues([0, 1, 2, 3]), f'x_{i}_{d}')

                # MATRIKS INDIKATOR STATUS (OFF & SHIFT 3)
                is_off_matrix = {}
                is_s3_matrix = {}
                for i in range(num_karyawan):
                    for d in range(num_hari):
                        is_off = model.NewBoolVar(f'is_off_{i}_{d}')
                        model.Add(x[i, d] == 0).OnlyEnforceIf(is_off)
                        model.Add(x[i, d] != 0).OnlyEnforceIf(is_off.Not())
                        is_off_matrix[i, d] = is_off

                        is_s3 = model.NewBoolVar(f'is_s3_{i}_{d}')
                        model.Add(x[i, d] == 3).OnlyEnforceIf(is_s3)
                        model.Add(x[i, d] != 3).OnlyEnforceIf(is_s3.Not())
                        is_s3_matrix[i, d] = is_s3

                # ATURAN 1: Jatah Libur Bulanan Mutlak (28-30 hari = 8 OFF, 31 hari = 9 OFF)
                total_off_wajib = 8 if num_hari in [28, 29, 30] else 9
                for i in range(num_karyawan):
                    model.Add(sum(is_off_matrix[i, d] for d in range(num_hari)) == total_off_wajib)

                # ATURAN KETAT: SELESAI SHIFT 3 SELAMA 3 HARI BERTURUT-TURUT WAJIB LIBUR (OFF)
                for i in range(num_karyawan):
                    full_s3_timeline = [1 if c == 3 else 0 for c in riwayat_mei_seminggu[i][-3:]] + [is_s3_matrix[i, d] for d in range(num_hari)]
                    full_off_timeline = [1 if c == 0 else 0 for c in riwayat_mei_seminggu[i][-3:]] + [is_off_matrix[i, d] for d in range(num_hari)]
                    
                    for start_idx in range(len(full_s3_timeline) - 3):
                        day1_s3 = full_s3_timeline[start_idx]
                        day2_s3 = full_s3_timeline[start_idx + 1]
                        day3_s3 = full_s3_timeline[start_idx + 2]
                        day4_off = full_off_timeline[start_idx + 3]
                        
                        trigger = model.NewBoolVar(f's3_3d_trig_{i}_{start_idx}')
                        model.AddBoolAnd([day1_s3, day2_s3, day3_s3]).OnlyEnforceIf(trigger)
                        model.AddBoolOr([day1_s3.Not(), day2_s3.Not(), day3_s3.Not()]).OnlyEnforceIf(trigger.Not())
                        model.Add(day4_off == 1).OnlyEnforceIf(trigger)

                # OPTIMASI PRIORITAS BERBOBOT (SOFT CONSTRAINTS)
                # Target A: Sebisa mungkin hari libur digandengkan berturut-turut
                consec_off_vars = []
                for i in range(num_karyawan):
                    for d in range(num_hari - 1):
                        pair_consec = model.NewBoolVar(f'pair_consec_{i}_{d}')
                        model.AddBoolAnd([is_off_matrix[i, d], is_off_matrix[i, d+1]]).OnlyEnforceIf(pair_consec)
                        model.AddBoolOr([is_off_matrix[i, d].Not(), is_off_matrix[i, d+1].Not()]).OnlyEnforceIf(pair_consec.Not())
                        consec_off_vars.append(pair_consec)

                # Target B: Mengusahakan 2 Hari Libur per Minggu Penuh Kalender
                week_per_emp_ideal = []
                saut_friday_s3_vars = [] 
                riwayat_off_bulan_lalu = {i: [1 if s == 0 else 0 for s in riwayat_mei_seminggu[i]] for i in range(num_karyawan)}

                weeks = []
                current_week = []
                for d in range(num_hari):
                    current_week.append(d)
                    if hari_ke_nama[d] == 'Minggu' or d == num_hari - 1:
                        weeks.append(current_week)
                        current_week = []

                for idx_w, week in enumerate(weeks):
                    if idx_w == 0 and hari_pertama_idx > 0:
                        for i in range(num_karyawan):
                            total_off_prev = sum(riwayat_off_bulan_lalu[i][-hari_pertama_idx:])
                            total_off_this_week = total_off_prev + sum(is_off_matrix[i, d] for d in week)
                            model.Add(total_off_this_week >= 1)
                            model.Add(total_off_this_week <= 3)
                            
                            is_ideal = model.NewBoolVar(f'ideal_w0_i_{i}')
                            model.Add(total_off_this_week == 2).OnlyEnforceIf(is_ideal)
                            model.Add(total_off_this_week != 2).OnlyEnforceIf(is_ideal.Not())
                            week_per_emp_ideal.append(is_ideal)
                    elif len(week) == 7:
                        for i in range(num_karyawan):
                            total_off_this_week = sum(is_off_matrix[i, d] for d in week)
                            model.Add(total_off_this_week >= 1)
                            model.Add(total_off_this_week <= 3)
                            
                            is_ideal = model.NewBoolVar(f'ideal_w{idx_w}_i_{i}')
                            model.Add(total_off_this_week == 2).OnlyEnforceIf(is_ideal)
                            model.Add(total_off_this_week != 2).OnlyEnforceIf(is_ideal.Not())
                            week_per_emp_ideal.append(is_ideal)
                    else:
                        for i in range(num_karyawan):
                            model.Add(sum(is_off_matrix[i, d] for d in week) >= 0)
                            model.Add(sum(is_off_matrix[i, d] for d in week) <= 2)

                if saut_idx != -1:
                    for d in range(num_hari):
                        if hari_ke_nama[d] == 'Jumat':
                            saut_friday_s3_vars.append(is_s3_matrix[saut_idx, d])

                model.Maximize(10 * sum(week_per_emp_ideal) + 5 * sum(consec_off_vars) + 15 * sum(saut_friday_s3_vars))

                # =====================================================================
                # ATURAN OPERASIONAL SHIFT DAN REGULASI KARYAWAN
                # =====================================================================
                for d in range(num_hari):
                    if saut_idx != -1:
                        model.Add(x[saut_idx, d] != 2) # Saut bebas dari Shift 2

                    for s in [1, 2, 3]:
                        is_in_shift = []
                        for i in range(num_karyawan):
                            in_shift = model.NewBoolVar(f's_{s}_i_{i}_d_{d}')
                            model.Add(x[i, d] == s).OnlyEnforceIf(in_shift)
                            model.Add(x[i, d] != s).OnlyEnforceIf(in_shift.Not())
                            is_in_shift.append(in_shift)
                        
                        if s == 3:
                            model.Add(sum(is_in_shift) == 2) 
                        else:
                            model.Add(sum(is_in_shift) >= 1)
                            model.Add(sum(is_in_shift) <= 3)

                        if saut_idx != -1:
                            saut_disini = model.NewBoolVar(f'saut_is_at_s_{s}_d_{d}')
                            model.Add(x[saut_idx, d] == s).OnlyEnforceIf(saut_disini)
                            model.Add(x[saut_idx, d] != s).OnlyEnforceIf(saut_disini.Not())
                            model.Add(sum(is_in_shift) == 2).OnlyEnforceIf(saut_disini)

                # LOGIKA PROPORSI SHIFT 1 SAUT (MINIMAL 12 - MAKSIMAL 18 HARI)
                for i in range(num_karyawan):
                    emp_s1_days = []
                    for d in range(num_hari):
                        is_s1 = model.NewBoolVar(f's1_count_{i}_{d}')
                        model.Add(x[i, d] == 1).OnlyEnforceIf(is_s1)
                        model.Add(x[i, d] != 1).OnlyEnforceIf(is_s1.Not())
                        emp_s1_days.append(is_s1)
                    
                    if i == saut_idx and saut_idx != -1:
                        model.Add(sum(emp_s1_days) >= 12) 
                        model.Add(sum(emp_s1_days) <= 18) 
                    else:
                        model.Add(sum(emp_s1_days) >= 1)
                        model.Add(sum(emp_s1_days) <= 10)

                # MAKSIMAL 5 HARI KERJA BERTURUT-TURUT LINTAS BATAS BULAN
                for i in range(num_karyawan):
                    prev_work_offs = [1 if s == 0 else 0 for s in riwayat_mei_seminggu[i][-5:]]
                    current_work_offs = [is_off_matrix[i, d] for d in range(num_hari)]
                    all_work_offs = prev_work_offs + current_work_offs
                    for start_day in range(len(all_work_offs) - 5):
                        window = all_work_offs[start_day : start_day + 6]
                        model.Add(sum(window) >= 1)

                # TRANSISI ISTIRAHAT MINIMAL PASCA SHIFT 3
                for i in range(num_karyawan):
                    last_mei = riwayat_mei_seminggu[i][-1]
                    if last_mei == 3:
                        model.AddAllowedAssignments([x[i, 0]], [[0], [3]])
                    elif last_mei == 2:
                        model.Add(x[i, 0] != 1)
                        
                    for d in range(num_hari - 1):
                        model.AddAllowedAssignments([x[i, d], x[i, d+1]], [[0,0],[0,1],[0,2],[0,3],[1,0],[1,1],[1,2],[1,3],[2,0],[2,2],[2,3],[3,0],[3,3]])

                # DISTRIBUSI ADIL SHIFT 3 (MALAM)
                for i in range(num_karyawan):
                    if i != 0:
                        emp_s3_days = []
                        for d in range(num_hari):
                            is_s3 = model.NewBoolVar(f'fair_s3_{i}_{d}')
                            model.Add(x[i, d] == 3).OnlyEnforceIf(is_s3)
                            model.Add(x[i, d] != 3).OnlyEnforceIf(is_s3.Not())
                            emp_s3_days.append(is_s3)
                        
                        if i == saut_idx:
                            model.Add(sum(emp_s3_days) >= 1)
                            model.Add(sum(emp_s3_days) <= 5)
                        else:
                            model.Add(sum(emp_s3_days) >= 8)
                            model.Add(sum(emp_s3_days) <= 14)

                # Eksekusi Solver
                solver = cp_model.CpSolver()
                solver.parameters.linearization_level = 0
                solver.parameters.max_time_in_seconds = 20.0
                status = solver.Solve(model)

                if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Sheet1"
                    ws.views.sheetView[0].showGridLines = True

                    fill_s1 = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    font_s1 = Font(name="Calibri", size=11, color="006100", bold=True)
                    fill_s2 = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    font_s2 = Font(name="Calibri", size=11, color="9C6500", bold=True)
                    fill_s3 = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
                    font_s3 = Font(name="Calibri", size=11, color="1F4E78", bold=True)
                    fill_off = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    font_off = Font(name="Calibri", size=11, color="A6A6A6")
                    font_header = Font(name="Calibri", size=11, bold=True)
                    fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    thin_border = Border(
                        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
                    )

                    ws.cell(row=1, column=1, value=date_bulan_depan.strftime("%Y-%m-%d")).font = font_header

                    headers_kiri = ["Nama", "Layer", "Shift 1", "Shift 2", "Shift 3", "Libur", "Kerja", ""]
                    for col_idx, text in enumerate(headers_kiri, start=1):
                        cell = ws.cell(row=2, column=col_idx, value=text)
                        cell.font = font_header; cell.fill = fill_header

                    for col_idx, text in enumerate(["Shift 1", "Shift 2", "Shift 3", "Libur", "Kerja"], start=3):
                        ws.cell(row=3, column=col_idx, value=text).font = font_header

                    for r in [2, 3]:
                        for c in range(1, 9):
                            ws.cell(row=r, column=c).fill = fill_header
                            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center" if c > 1 else "left", vertical="center")

                    col_mapping = {}
                    current_col = 9
                    for d in range(num_hari):
                        col_mapping[d] = current_col
                        ws.cell(row=2, column=current_col, value=d+1).font = font_header
                        ws.cell(row=2, column=current_col).fill = fill_header
                        ws.cell(row=2, column=current_col).alignment = Alignment(horizontal="center")
                        
                        ws.cell(row=3, column=current_col, value=hari_ke_nama[d]).font = font_header
                        ws.cell(row=3, column=current_col).fill = fill_header
                        ws.cell(row=3, column=current_col).alignment = Alignment(horizontal="center")
                        
                        if hari_ke_nama[d] == 'Minggu' and d != num_hari - 1:
                            current_col += 1
                            ws.cell(row=2, column=current_col).fill = fill_header
                            ws.cell(row=3, column=current_col).fill = fill_header
                        current_col += 1

                    for i in range(num_karyawan):
                        row_idx = i + 4
                        s1_c = sum(solver.Value(x[i, d]) == 1 for d in range(num_hari))
                        s2_c = sum(solver.Value(x[i, d]) == 2 for d in range(num_hari))
                        s3_c = sum(solver.Value(x[i, d]) == 3 for d in range(num_hari))
                        off_c = sum(solver.Value(x[i, d]) == 0 for d in range(num_hari))
                        kerja_c = s1_c + s2_c + s3_c
                        
                        ws.cell(row=row_idx, column=1, value=karyawan[i]).font = Font(name="Calibri", size=11)
                        ws.cell(row=row_idx, column=2, value="" if i == 0 else "L1").alignment = Alignment(horizontal="center")
                        ws.cell(row=row_idx, column=3, value=s1_c).alignment = Alignment(horizontal="center")
                        ws.cell(row=row_idx, column=4, value=s2_c).alignment = Alignment(horizontal="center")
                        ws.cell(row=row_idx, column=5, value=s3_c).alignment = Alignment(horizontal="center")
                        ws.cell(row=row_idx, column=6, value=off_c).alignment = Alignment(horizontal="center")
                        ws.cell(row=row_idx, column=7, value=kerja_c).alignment = Alignment(horizontal="center")
                        
                        for c in range(1, 9): ws.cell(row=row_idx, column=c).border = thin_border
                            
                        for d in range(num_hari):
                            target_col = col_mapping[d]
                            val = solver.Value(x[i, d])
                            cell = ws.cell(row=row_idx, column=target_col)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            
                            if val == 1:
                                cell.value = 1; cell.fill = fill_s1; cell.font = font_s1
                            elif val == 2:
                                cell.value = 2; cell.fill = fill_s2; cell.font = font_s2
                            elif val == 3:
                                cell.value = 3; cell.fill = fill_s3; cell.font = font_s3
                            else:
                                cell.value = "OFF"; cell.fill = fill_off; cell.font = font_off

                    # --- METADATA TABEL BAWAH ---
                    start_row_meta = 14  
                    
                    # TABEL 1: Referensi Shift Waktu (Kolom A-D)
                    meta_shift_headers = ["Shift", "Waktu Masuk", "Waktu Pulang", "Status"]
                    meta_shift_data = [
                        ["1", "07.00", "16.00", "Kerja"],
                        ["2", "14.00", "23.00", "Kerja"],
                        ["3", "22.30", "07.30", "Kerja"],
                        ["OFF", " ", " ", "Libur"],  
                        ["CUTI", " ", " ", "Cuti"]   
                    ]
                    
                    for col_idx, h_text in enumerate(meta_shift_headers, start=1):
                        cell = ws.cell(row=start_row_meta, column=col_idx, value=h_text)
                        cell.font = font_header; cell.fill = fill_header; cell.alignment = Alignment(horizontal="center"); cell.border = thin_border
                        
                    for r_idx, row_content in enumerate(meta_shift_data, start=start_row_meta + 1):
                        for c_idx, val in enumerate(row_content, start=1):
                            cell = ws.cell(row=r_idx, column=c_idx, value=val)
                            cell.font = Font(name="Calibri", size=11); cell.border = thin_border; cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left")

                    # TABEL 2: Daftar Email Agent Independen (Kolom G-H)
                    start_col_email = 7  
                    meta_email_headers = ["Nama", "Email"]
                    meta_email_data = [
                        ["Jasmine Al-Rosamund", "jasmine.rosamund@alto.id"],
                        ["Reyonal Novrianto", "reyonal@alto.id"],
                        ["Eko Wahyudi", "eko.wahuydi@alto.id"],
                        ["Chairul Anwar", "chairul.anwar@alto.id"],
                        ["Taufan Maulana", "taufan.maulana@alto.id"],
                        ["Faizal", "Faizal@alto.id"],
                        ["Saut Parsaulian", "saut@alto.id"]
                    ]
                    
                    for col_idx, h_text in enumerate(meta_email_headers, start=start_col_email):
                        cell = ws.cell(row=start_row_meta, column=col_idx, value=h_text)
                        cell.font = font_header; cell.fill = fill_header; cell.border = thin_border
                        cell.alignment = Alignment(horizontal="center" if col_idx > start_col_email else "left")
                        
                    for email_r_idx, email_row_content in enumerate(meta_email_data, start=start_row_meta + 1):
                        cell_nama = ws.cell(row=email_r_idx, column=start_col_email, value=email_row_content[0])
                        cell_nama.font = Font(name="Calibri", size=11); cell_nama.border = thin_border; cell_nama.alignment = Alignment(horizontal="left")
                        
                        cell_mail = ws.cell(row=email_r_idx, column=start_col_email + 1, value=email_row_content[1])
                        cell_mail.font = Font(name="Calibri", size=11); cell_mail.border = thin_border; cell_mail.alignment = Alignment(horizontal="left")

                    for col in ws.columns:
                        max_len = max(len(str(cell.value or '')) for cell in col)
                        col_letter = get_column_letter(col[0].column)
                        ws.column_dimensions[col_letter].width = max(max_len + 3, 6)
                    ws.column_dimensions['A'].width = 38
                    ws.column_dimensions['G'].width = 38  

                    excel_buffer = io.BytesIO()
                    wb.save(excel_buffer)
                    excel_buffer.seek(0)
                    
                    st.success("🎉 Penjadwalan sukses dibentuk tanpa ada aturan yang melanggar!")
                    st.download_button(
                        label="📥 Download Berkas Excel Jadwal Baru",
                        data=excel_buffer,
                        file_name=f"Jadwal_Otomatis_{date_bulan_depan.strftime('%B_%Y')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.error("Gagal! Algoritma mendeteksi adanya bentrokan aturan mutlak. Mohon periksa kembali kesesuaian jatah libur tim.")
    except Exception as e:
        st.error(f"Terjadi kesalahan format pembacaan file: {e}. Pastikan file template sesuai dengan format aslinya.")
